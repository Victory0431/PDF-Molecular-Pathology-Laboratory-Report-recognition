# -*- coding: UTF-8 -*-

import re
import os
import glob
import datetime
import pdfplumber
import openpyxl
from openpyxl import load_workbook

def userinfo(first_table,each_pdfinfo):
    each_pdfinfo['user_name'] = first_table[0][1][1]
    each_pdfinfo['user_gender'] = first_table[0][2][1]
    each_pdfinfo['sample_id'] = first_table[0][2][5]
    each_pdfinfo['user_age'] = first_table[0][3][1]
    each_pdfinfo['sample_sorce'] = first_table[0][4][5]
    each_pdfinfo['diagnose'] = first_table[0][6][1]
    each_pdfinfo['sample_date'] = first_table[0][6][5]
    each_pdfinfo['project'] = first_table[0][7][1]
    each_pdfinfo['docter'] = first_table[0][5][1]
    each_pdfinfo['test_date'] = first_table[0][7][5]

def Parsing_type(input_txt):
    each_gene_type = []
    for each_line in input_txt.split("\n"):
        if ("以下基因检测到" in each_line):
            matchObj  = re.match( r'.*?变异\((.*?)\)', each_line , re.M|re.I)
            each_gene_type.append(matchObj.group(1))
    return each_gene_type

def Parsing_userinfo(input_txt,info_dict):
    # each_gene_type = []
    for each_line in input_txt.split("\n"):
        if ("姓 名" in each_line):
            matchObj  = re.match( r'姓 名 (.*?) 送检医院.*', each_line , re.M|re.I)
            try:
                info_dict['user_name'] = matchObj.group(1)
            except AttributeError:
                info_dict['user_name'] = ''

        if ("性 别" in each_line):
            matchObj  = re.match( r'性 别 (.*?)科 室 血液科 标本条码 (.*)', each_line , re.M|re.I)
            try:
                info_dict['user_gender'] = matchObj.group(1)
            except AttributeError:
                info_dict['user_gender'] = '' 

            try:
                info_dict['sample_id'] = matchObj.group(2)
            except:
                info_dict['sample_id'] = ''
        
        if ("年 龄" in each_line):
            matchObj  = re.match( r'年 龄 (.*?)门诊/住院号(.*)', each_line , re.M|re.I)
            try:
                info_dict['user_age'] = matchObj.group(1)
            except AttributeError:
                info_dict['user_age'] = ''
        
        if ("标本类型" in each_line):
            matchObj  = re.match( r'.*?标本类型(.*)', each_line , re.M|re.I)
            try:
                info_dict['sample_sorce'] = matchObj.group(1)
            except AttributeError:
                info_dict['sample_sorce'] = ''

        
        if ("申请医生" in each_line):
            matchObj  = re.match( r'申请医生 (.*?)医生电话(.*)', each_line , re.M|re.I)
            try:
                info_dict['docter'] = matchObj.group(1)
            except AttributeError:
                info_dict['docter'] = ''
        if ("临床诊断" in each_line):
            matchObj  = re.match( r'临床诊断(.*?)医院标识.*?采样时间(.*)', each_line , re.M|re.I)
            try:
                info_dict['diagnose'] = matchObj.group(1)
            except AttributeError:
                info_dict['diagnose'] = ''
            try:
                info_dict['sample_date'] = matchObj.group(2)
            except AttributeError:
                info_dict['sample_date'] = ''
        
        if ("项目名称" in each_line):
            matchObj  = re.match( r'项目名称 (.*?) 接收时间 (.*)', each_line , re.M|re.I)
            try:
                info_dict['project'] = matchObj.group(1)
            except AttributeError:
                info_dict['project'] = ''
            try:
                info_dict['test_date'] = matchObj.group(2)
            except AttributeError:
                info_dict['test_date'] = ''

if os.path.exists('输出模板.xlsx'):
    wb6 = load_workbook('输出模板.xlsx')
    ws6 = wb6.active
    
else:
    wb = openpyxl.Workbook()
    wb.save('输出模板.xlsx')
    wb6 = load_workbook('输出模板.xlsx')
    ws6 = wb6.active
    ws6['A' + str(1)] = '编号'
    ws6['B' + str(1)] = '姓名'
    ws6['C' + str(1)] = '性别'
    ws6['D' + str(1)] = '年龄'
    ws6['E' + str(1)] = '标本采集日期'
    ws6['F' + str(1)] = '送检日期'
    ws6['G' + str(1)] = '标本来源'
    ws6['H' + str(1)] = '初始诊断'
    ws6['I' + str(1)] = '检测项目'
    ws6['J' + str(1)] = '送检医师'
    ws6['K' + str(1)] = '标本条码'
    ws6['L' + str(1)] = 'TierI_突变基因'
    ws6['M' + str(1)] = 'TierI_突变命名'
    ws6['N' + str(1)] = 'TierI_突变频率'
    ws6['O' + str(1)] = 'TierII_突变基因'
    ws6['P' + str(1)] = 'TierII_突变命名'
    ws6['Q' + str(1)] = 'TierII_突变频率'
    ws6['R' + str(1)] = 'TierIII_突变基因'
    ws6['S' + str(1)] = 'TierIII_突变命名'
    ws6['T' + str(1)] = 'TierIII_突变频率'
    wb6.save('输出模板.xlsx')
path = os.getcwd()

wb6 = load_workbook('输出模板.xlsx')
ws6 = wb6.active
line = 2
numuser = 1

pdf_flist = glob.glob(path + "\\*pdf")
# print(pdf_flist)
pdf_number = 1
all_info = {}
for each_pdf in pdf_flist:
    print(each_pdf)
    each_pdfinfo = {
        'user_name'   : '', # 姓名
        'user_gender' : '', # 性别
        'user_age'    : '', # 年龄
        'sample_date' : '', # 标本采集日期
        'test_date'   : '', # 送检日期
        'sample_sorce': '', # 标本来源
        'diagnose'    : '', # 初始诊断
        'project'     : '', # 检测项目
        'docter'      : '', # 送检医师
        'sample_id'   : '', # 标本条码
        'Tier I'      : [],
        'Tier II'     : [],
        'Tier III'    : [],
    }
    with pdfplumber.open(each_pdf) as pdf:
        #基本信息定义获取
        first_page = pdf.pages[0]
        second_page = pdf.pages[1]
        #third_page = pdf.pages[2]
        page01 = first_page.extract_text()
        page02 = second_page.extract_text()
        gene_type01 = Parsing_type(page01)
        gene_type02 = Parsing_type(page02)
        first_table = first_page.extract_tables()
        print(gene_type01)
        print(len(first_table))
        second_table = second_page.extract_tables()
        print(gene_type02)
        print(len(second_table))
        #third_table = third_page.extract_tables()
        
    flag = 0
    #userinfo获取
    try:
        userinfo(first_table,each_pdfinfo)
    except:
        Parsing_userinfo(page01,each_pdfinfo)
        flag = 1
    #userinfo写入
    print(each_pdfinfo)
    ws6['A' + str(line)] = numuser
    ws6['B' + str(line)] = each_pdfinfo['user_name']
    ws6['C' + str(line)] = each_pdfinfo['user_gender']
    ws6['D' + str(line)] = each_pdfinfo['user_age']
    ws6['E' + str(line)] = each_pdfinfo['sample_date']
    ws6['F' + str(line)] = each_pdfinfo['test_date']
    ws6['G' + str(line)] = each_pdfinfo['sample_sorce']
    ws6['H' + str(line)] = each_pdfinfo['diagnose']
    ws6['I' + str(line)] = each_pdfinfo['project']
    ws6['J' + str(line)] = each_pdfinfo['docter']
    ws6['K' + str(line)] = each_pdfinfo['sample_id']

    #开始处理检测结果Tier I - II - III
    result = {}
    if flag: #第一个用户表格识别不出情况
        for i in range(len(gene_type01)):
            result[gene_type01[i]] = first_table[i][1:]
            
        for i in range(len(gene_type02)):
            result[gene_type02[i]] = second_table[i][1:]
    else:
        for i in range(len(gene_type01)):
            result[gene_type01[i]] = first_table[i+1][1:]
            
        for i in range(len(gene_type02)):
            result[gene_type02[i]] = second_table[i+1][1:]

    #Tier I - II - III 写入
    line1 = line
    line2 = line
    line3 = line
    for i in result:
        if i == 'Tier I': 
            tier1 = result[i]
            for j in tier1:
                ws6['L' + str(line1)] = j[0]
                ws6['M' + str(line1)] = j[1]
                ws6['N' + str(line1)] = j[2]
                line1 += 1
        elif i == 'Tier II':
            tier2 = result[i]
            for j in tier2:
                ws6['O' + str(line2)] = j[0]
                ws6['P' + str(line2)] = j[1]
                ws6['Q' + str(line2)] = j[2]
                line2 += 1
        elif i == 'Tier III':
            tier3 = result[i]
            for j in tier3:
                ws6['R' + str(line3)] = j[0]
                ws6['S' + str(line3)] = j[1]
                ws6['T' + str(line3)] = j[2]
                line3 += 1
                
    realine = max(line1,line2,line3)-1
    print(result)   
    line = realine + 1
    numuser += 1

t1 = datetime.datetime.now()
st1 = str(t1)
wb6.save(st1[:-7].replace(':','-') + '.xlsx')   


